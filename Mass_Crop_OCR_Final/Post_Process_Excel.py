import openpyxl
import re
from openpyxl.utils import column_index_from_string

def rearrange_columns(worksheet, row, arrangement):
    print("Rearranging row:", row[0].row)  
    print("Arrangement:", arrangement) 

    new_values = []
    for col_letter in arrangement:
        source_col = worksheet.cell(row=row[0].row, column=column_index_from_string(col_letter)).value
        print("Extracted value:", source_col)  
        new_values.append(source_col)

    print("Values to write:", new_values) 
    for col_idx, val in enumerate(new_values, start=1): 
        worksheet.cell(row=row[0].row, column=col_idx).value = val

def split_and_insert_numbers(worksheet, column_to_check='E'):
    """Shifts columns after E to the right, then splits and inserts numbers.

    Args:
        worksheet: The worksheet to process.
        column_to_check: The column to check for 3-digit numbers (default: 'E').
    """

    column_index = column_index_from_string(column_to_check)

    for row in worksheet.iter_rows(): 
        occupied_cells = 0
        for cell in row:
            if cell.value is not None:
                occupied_cells += 1

        if occupied_cells == 8:  
            # 1. Shift existing data in the current row
            for col in range(worksheet.max_column, column_index, -1):
                worksheet.cell(row=row[0].row, column=col + 2).value = worksheet.cell(row=row[0].row, column=col).value

            # 2. Process cell in column E of the current row
            cell = worksheet[column_to_check + str(row[0].row)]

            if cell.value:
                match = re.search(r'\d{3}(?:[\s\W]+\d{3}){2}', str(cell.value))
                if match:
                    numbers = match.group(0).split()

                    # Insert with bounds checking 
                    for i, number in enumerate(numbers[:3]): 
                        clean_number = re.sub(r"\D", "", number)
                        if clean_number: 
                            worksheet.cell(row=row[0].row, column=column_index + i).value = int(clean_number)
        else:
            if occupied_cells == 13:
                pass
            else:
             for col in range(2, worksheet.max_column + 1):  
                worksheet.cell(row=row[0].row, column=col).value = None

    for row in worksheet.iter_rows(): 
        occupied_cells = 0
        contains_text = False 

        for cell in row[1:]:  # Start from the second cell (skip the first)
            if cell.value is not None:
                occupied_cells += 1
                if not str(cell.value).isdigit():
                    contains_text = True
                    
            if contains_text:
                for col in range(2, worksheet.max_column + 1):  
                 worksheet.cell(row=row[0].row, column=col).value = None        

def process_excel_after_split(worksheet):
    for row in worksheet.iter_rows(): 
        occupied_cells = sum(cell.value is not None for cell in row)  # Recount cells

        if occupied_cells == 13:
            if 'L' in row[0].value:   
                rearrange_columns(worksheet, row, ['A','B', 'C', 'E', 'G', 'I', 'K', 'M', 'L', 'J', 'H', 'F', 'D']) 
            elif 'R' in row[0].value:
                rearrange_columns(worksheet, row, ['A','B', 'D', 'F', 'H', 'J', 'L', 'M', 'K', 'I', 'G', 'E', 'C'])

        elif occupied_cells == 10:  # Check for 10 columns after modification
            if 'L' in row[0].value:
                rearrange_columns(worksheet, row, ['A','B', 'D', 'J', 'H', 'C', 'E', 'I', 'F'])
            elif 'R' in row[0].value:
                rearrange_columns(worksheet, row, ['A','B', 'H', 'J', 'D', 'C', 'G', 'I', 'E', 'F'])  
         
    
# Load your workbook
workbook = openpyxl.load_workbook("/Users/mertegemencaliskan/Downloads/Output_OCR/ocr_results.xlsx")
worksheet = workbook["Sheet"] 

# Specify the column to check (default: 'E')
column_to_check = 'E' 

# Split and insert numbers
split_and_insert_numbers(worksheet, column_to_check)
process_excel_after_split(worksheet) # Then, rearrange columns

# Save the workbook
workbook.save("ready_file.xlsx")