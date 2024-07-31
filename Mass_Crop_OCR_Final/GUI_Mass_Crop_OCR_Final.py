import sys
import os
import openpyxl
import cv2
import numpy as np
import easyocr
from PyQt5.QtWidgets import QApplication, QProgressBar, QWidget, QPushButton, QLabel, QFileDialog, QLineEdit, QMessageBox, QVBoxLayout
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtCore import Qt
import openpyxl
import re
from openpyxl.utils import column_index_from_string

class ImageCropper(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("OCT 3D Angio & Macula Image Cropper & OCR by BioStemCore ")
        self.setGeometry(100, 100, 1280, 720)

        self.cropping_areas = {}  

        # Labels
        
        self.ocr_output_path_label = QLabel("OCR Results Output Folder:", self)
        self.ocr_output_path_label.setGeometry(20, 10, 200, 20)
        
        self.ocr_output_path_text = QLineEdit(self)
        self.ocr_output_path_text.setGeometry(220, 10, 1040, 20) 
       
        self.browse_ocr_output_button = QPushButton("Browse", self)
        self.browse_ocr_output_button.setGeometry(10, 40, 1260, 30)
        self.browse_ocr_output_button.clicked.connect(self.browse_ocr_output_folder)
       
        self.ocr_crop_button = QPushButton("Run OCR on Crop Output Folder", self)
        self.ocr_crop_button.setGeometry(10, 80, 1260, 30)
        self.ocr_crop_button.clicked.connect(self.run_ocr_on_crop_output_folder) 
        
        self.output_path_label = QLabel("Crop Output Folder:", self)
        self.output_path_label.setGeometry(10, 530, 120, 20)

        self.output_path_text = QLineEdit(self)
        self.output_path_text.setGeometry(150, 530, 800, 20)

        self.browse_output_button = QPushButton("Browse", self)
        self.browse_output_button.setGeometry(1000, 525, 100, 30)
        self.browse_output_button.clicked.connect(self.browse_output_folder)

        self.input_path_label = QLabel("Crop Input Folder:", self)
        self.input_path_label.setGeometry(10, 560, 120, 20)

        self.input_path_text = QLineEdit(self)
        self.input_path_text.setGeometry(150, 560, 800, 20)

        self.browse_input_button = QPushButton("Browse", self)
        self.browse_input_button.setGeometry(1000, 555, 100, 30)
        self.browse_input_button.clicked.connect(self.browse_input_folder)

        self.crop_button = QPushButton("Crop and Save", self)
        self.crop_button.setGeometry(1110, 525, 120, 60)
        self.crop_button.clicked.connect(self.crop_and_save)

        # Progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setGeometry(10, 590, 1260, 20)



    def run_ocr_on_crop_output_folder(self):
        
        ocr_output_path = self.ocr_output_path_text.text()
        output_folder = self.output_path_text.text()
        
        
        if self.ocr_output_path_text.text() == "":
         QMessageBox.warning(self, "Warning", "Please select an output folder for OCR results.")
         return
    
        if self.output_path_text.text() == "":
         QMessageBox.warning(self, "Warning", "Please select output cropped images folder as input folder to run OCR on.")
         return
         # Initialize EasyOCR reader
      
        reader = easyocr.Reader(['en']) # Assuming English language

        # Prepare Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Process images in the output folder
        image_files = [f for f in os.listdir(output_folder) if f.endswith(('.jpg', '.jpeg', '.png', '.bmp'))]

        for index, image_file in enumerate(image_files):
          image_path = os.path.join(output_folder, image_file)

          # Perform OCR
          result = reader.readtext(image_path, detail=0)

          # Write to Excel
          sheet.cell(row=index + 1, column=1).value = image_file
          column_index = 2  # Start writing OCR words in the second column
          for word in result:
            sheet.cell(row=index + 1, column=column_index).value = word
            column_index += 1

        default_filename = "ocr_results.xlsx"  # You can customize the default name
        output_file = os.path.join(ocr_output_path, default_filename)
        workbook.save(output_file)
        
        def rearrange_columns(worksheet, row, arrangement):
            print("Rearranging row:", row[0].row)  
            print("Arrangement:", arrangement) 

            new_values = []
            for col_letter in arrangement:
                source_col = worksheet.cell(row=row[0].row, column=column_index_from_string(col_letter)).value
                print("Extracted value:", source_col)  
                new_values.append(source_col)

            print("Values to write:", new_values) 
            for col_idx, val in enumerate(new_values, start=1): 
                worksheet.cell(row=row[0].row, column=col_idx).value = val

        def split_and_insert_numbers(worksheet, column_to_check='E'):
            """Shifts columns after E to the right, then splits and inserts numbers.

            Args:
                worksheet: The worksheet to process.
                column_to_check: The column to check for 3-digit numbers (default: 'E').
            """

            column_index = column_index_from_string(column_to_check)

            for row in worksheet.iter_rows(): 
                occupied_cells = 0
                for cell in row:
                    if cell.value is not None:
                        occupied_cells += 1

                if occupied_cells == 8:  
                    # 1. Shift existing data in the current row
                    for col in range(worksheet.max_column, column_index, -1):
                        worksheet.cell(row=row[0].row, column=col + 2).value = worksheet.cell(row=row[0].row, column=col).value

                    # 2. Process cell in column E of the current row
                    cell = worksheet[column_to_check + str(row[0].row)]

                    if cell.value:
                        match = re.search(r'\d{3}(?:[\s\W]+\d{3}){2}', str(cell.value))
                        if match:
                            numbers = match.group(0).split()

                            # Insert with bounds checking 
                            for i, number in enumerate(numbers[:3]): 
                                clean_number = re.sub(r"\D", "", number)
                                if clean_number: 
                                    worksheet.cell(row=row[0].row, column=column_index + i).value = int(clean_number)
                else:
                    if occupied_cells == 13:
                        pass
                    else:
                     for col in range(2, worksheet.max_column + 1):  
                        worksheet.cell(row=row[0].row, column=col).value = None

            for row in worksheet.iter_rows(): 
                occupied_cells = 0
                contains_text = False 

                for cell in row[1:]:  # Start from the second cell (skip the first)
                    if cell.value is not None:
                        occupied_cells += 1
                        if not str(cell.value).isdigit():
                            contains_text = True
                            
                    if contains_text:
                        for col in range(2, worksheet.max_column + 1):  
                         worksheet.cell(row=row[0].row, column=col).value = None        

        def process_excel_after_split(worksheet):
            for row in worksheet.iter_rows(): 
                occupied_cells = sum(cell.value is not None for cell in row)  # Recount cells

                if occupied_cells == 13:
                    if 'L' in row[0].value:   
                        rearrange_columns(worksheet, row, ['A','B', 'C', 'E', 'G', 'I', 'K', 'M', 'L', 'J', 'H', 'F', 'D']) 
                    elif 'R' in row[0].value:
                        rearrange_columns(worksheet, row, ['A','B', 'D', 'F', 'H', 'J', 'L', 'M', 'K', 'I', 'G', 'E', 'C'])

                elif occupied_cells == 10:  # Check for 10 columns after modification
                    if 'L' in row[0].value:
                        rearrange_columns(worksheet, row, ['A','B', 'D', 'J', 'H', 'C', 'E', 'I', 'F'])
                    elif 'R' in row[0].value:
                        rearrange_columns(worksheet, row, ['A','B', 'H', 'J', 'D', 'C', 'G', 'I', 'E', 'F'])  
                
            
        # Load your workbook
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook["Sheet"] 

        # Specify the column to check (default: 'E')
        column_to_check = 'E' 

        # Split and insert numbers
        split_and_insert_numbers(worksheet, column_to_check)
        process_excel_after_split(worksheet) # Then, rearrange columns

        # Save the workbook
        workbook.save(output_file)
   
    def browse_ocr_output_folder(self):
        folder_dialog = QFileDialog.getExistingDirectory(self, "Select OCR Output Folder")
        self.ocr_output_path_text.setText(folder_dialog)

    def calculate_pixel_size(self, img):
       if img is None:
           return None  
       height, width = img.shape[:2]
       return (width, height)

    def browse_output_folder(self):
        folder_dialog = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        self.output_path_text.setText(folder_dialog)

    def browse_input_folder(self):
        folder_dialog = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        self.input_path_text.setText(folder_dialog)

    def crop_and_save(self):
    # Input/Output Validation
     if self.output_path_text.text() == "":
        QMessageBox.warning(self, "Warning", "Please select an output folder.")
        return
     if self.input_path_text.text() == "":
        QMessageBox.warning(self, "Warning", "Please select an input folder.")
        return

     input_folder = self.input_path_text.text()
     output_folder = self.output_path_text.text()

     # Step 1: Collect unique pixel sizes and get cropping coordinates
     pixel_sizes_and_coordinates = {}
     for root, dirs, files in os.walk(input_folder):
         if any(folder in root for folder in ["od_ad", "od_am", "os_am", "os_ad"]):
             continue

         for file_name in files:
             if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff','.bmp')) and 'screenshot' in file_name.lower():
                 file_path = os.path.join(root, file_name)
                 img = cv2.imread(file_path)
                 if img is None:
                     QMessageBox.warning(self, "Warning", f"Unable to read image: {file_path}")
                     continue

                 pixel_size = self.calculate_pixel_size(img)
                 if pixel_size not in pixel_sizes_and_coordinates:
                     # Get coordinates using OpenCV interactive process
                     self.run_opencv_code(file_path) 
                     if pixel_size in self.cropping_areas:  
                         pixel_sizes_and_coordinates[pixel_size] = self.cropping_areas[pixel_size]
                     else: 
                         QMessageBox.warning(self, "Warning", f"No cropping coordinates found for pixel size {pixel_size}")

     # Step 3: Crop all images
     total_files = 0
     for root, dirs, files in os.walk(input_folder):
         if any(folder in root for folder in ["od_ad", "od_am", "os_am", "os_ad"]):
             continue

         total_files += len(
             [f for f in files if f.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff','.bmp')) 
                         and 'screenshot' in f.lower()]
         )

     self.progress_bar.setMaximum(total_files)

     file_count = 0
     for root, dirs, files in os.walk(input_folder):
         if any(folder in root for folder in ["od_ad", "od_am", "os_am", "os_ad"]):
             continue

         for file_name in files:
             if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp')) and 'screenshot' in file_name.lower():
                 file_path = os.path.join(root, file_name)
                 img = cv2.imread(file_path)
                 if img is None:
                     QMessageBox.warning(self, "Warning", f"Unable to read image: {file_path}")
                     continue

                 pixel_size = self.calculate_pixel_size(img)
                 if pixel_size in pixel_sizes_and_coordinates:
                     x1, y1, x2, y2 = pixel_sizes_and_coordinates[pixel_size]
                     cropped_image = img[y1:y2, x1:x2]
                     output_path = os.path.join(output_folder, file_name)
                     cv2.imwrite(output_path, cropped_image) 

                 file_count += 1
                 self.progress_bar.setValue(file_count)
        
     QMessageBox.information(self, "Information", "Crop and save completed.")                  
                        
    def run_opencv_code(self, file_path):
        img = cv2.imread(file_path)
        if img is None:
          QMessageBox.warning(self, "Warning", f"Unable to read image: {file_path}")
          return 

        img_dup = np.copy(img)
        mouse_pressed = False
        starting_x = starting_y = ending_x = ending_y = -1
        pixel_size = self.calculate_pixel_size(img)

        def mousebutton(event, x, y, flags, param):
            nonlocal img_dup, starting_x, starting_y, ending_x, ending_y, mouse_pressed

            if event == cv2.EVENT_LBUTTONDOWN:
                mouse_pressed = True
                starting_x, starting_y = x, y
                img_dup = np.copy(img)

            elif event == cv2.EVENT_MOUSEMOVE:
                if mouse_pressed:
                    img_dup = np.copy(img)
                    cv2.rectangle(img_dup, (starting_x, starting_y), (x, y), (0, 255, 0), 1)

            elif event == cv2.EVENT_LBUTTONUP:
                mouse_pressed = False
                ending_x, ending_y = x, y

        cv2.namedWindow('image')
        cv2.setMouseCallback('image', mousebutton)

        while True:
            cv2.imshow('image', img_dup)
            k = cv2.waitKey(1)
            if k == ord('c'):
                if starting_y > ending_y:
                    starting_y, ending_y = ending_y, starting_y
                if starting_x > ending_x:
                    starting_x, ending_x = ending_x, starting_x
                if ending_y - starting_y > 1 and ending_x - starting_x > 0:
                    message_box = QMessageBox(self)
                    message_box.setIcon(QMessageBox.Question)
                    message_box.setText("Is this the area you want to select?")
                    message_box.setWindowTitle("Confirm Selection")
                    message_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                    message_box.setDefaultButton(QMessageBox.Yes)
                    if message_box.exec_() == QMessageBox.Yes:
                         # Update coordinates 
                         self.cropping_areas[pixel_size] = (starting_x, starting_y, ending_x, ending_y)
                         cv2.destroyAllWindows()
                         break 
                        
            elif k == 27:
                  cv2.destroyAllWindows()
                  break 
          
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ImageCropper()
    window.show()
    sys.exit(app.exec_())